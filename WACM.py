import pandas as pd
import numpy as np
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression, Ridge, ElasticNet, Lasso
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor, StackingRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from xgboost import XGBRegressor
from sklearn.metrics import r2_score
from sklearn.model_selection import GridSearchCV

def calculate_weighted_r2():
    file_list = [
        'Predicted_Well_B_ANN.xlsx', 'Predicted_Well_B_DT.xlsx', 
        'Predicted_Well_B_ElasticNetRegression.xlsx', 'Predicted_Well_B_GRNN.xlsx', 
        'Predicted_Well_B_KNN.xlsx', 'Predicted_Well_B_LassoRegression.xlsx', 
        'Predicted_Well_B_LightGBM.xlsx', 'Predicted_Well_B_LinearRegression.xlsx', 
        'Predicted_Well_B_LSTM.xlsx', 'Predicted_Well_B_PNN.xlsx', 
        'Predicted_Well_B_RBF.xlsx', 'Predicted_Well_B_RF.xlsx', 
        'Predicted_Well_B_RidgeRegression.xlsx', 'Predicted_Well_B_SVM.xlsx', 
        'Predicted_Well_B_XGBOOST.xlsx'
    ]
    
    y_actual = None
    y_predictions = []

    for file in file_list:
        data = pd.read_excel(file)
        if file == 'Predicted_Well_B_KNN.xlsx':
            y_ij = data.iloc[2:2486, 1].values
            y_pred_ij = data.iloc[2:2486, 2].values
        else:
            y_ij = data.iloc[2:2486, 4].values
            y_pred_ij = data.iloc[2:2486, 5].values
        
        if y_actual is None:
            y_actual = y_ij
        
        y_predictions.append(y_pred_ij)
    
    X = np.column_stack(y_predictions)
    y = y_actual

    estimators = [
        ('lstm', MLPRegressor()),  
        ('grnn', MLPRegressor()),  
        
        ('ann', MLPRegressor()),
        ('pnn', MLPRegressor()),  
        
        ('knn', KNeighborsRegressor()),  
        
        ('rbf', SVR(kernel='rbf')), 
        ('svm', SVR()),
        
        ('linreg', LinearRegression()),  
        ('ridge', Ridge()), 
        ('lasso', Lasso()),  
        ('enr', ElasticNet()),  
         
        ('lightgbm', GradientBoostingRegressor()), 
        ('rf', RandomForestRegressor()), 
        ('dt', DecisionTreeRegressor()), 
        ('xgboost', XGBRegressor())
    ]

    meta_model = Ridge()

    stacking_regressor = StackingRegressor(estimators=estimators, final_estimator=meta_model)
    
    param_grid = {
        'rf__n_estimators': [50, 100],
        'rf__max_depth': [10, 20]
    }

    grid_search = GridSearchCV(estimator=stacking_regressor, param_grid=param_grid, cv=3, n_jobs=-1, scoring='r2')
    grid_search.fit(X, y)

    best_model = grid_search.best_estimator_

    y_pred_combined = best_model.predict(X)
    weighted_r2 = r2_score(y, y_pred_combined)

    workbook = load_workbook("SA_CM.xlsx")
    sheet = workbook.active

    # Save weighted R²
    sheet.cell(row=17, column=5, value=weighted_r2)

    # Save individual R² scores
    individual_r2_scores = [r2_score(y, y_pred) for y_pred in y_predictions]
    for index, r2_value in enumerate(individual_r2_scores, start=2):
        sheet.cell(row=index, column=5, value=r2_value)

    # Save predicted vs actual data in column 16
    for i, y_pred in enumerate(y_predictions, start=2):
        for j, value in enumerate(y_pred, start=2):
            sheet.cell(row=j, column=16 , value=value)

    # Save combined predictions in column 16
    for i, value in enumerate(y_pred_combined, start=2):
        sheet.cell(row=i, column=16, value=value)

    workbook.save("SA_CM.xlsx")

    print(f"Final Weighted R²: {weighted_r2}")

calculate_weighted_r2()

print("Finished!")