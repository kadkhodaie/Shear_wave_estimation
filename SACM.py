import pandas as pd
from sklearn.preprocessing import StandardScaler , PolynomialFeatures
from sklearn.metrics import r2_score
from xgboost import XGBRegressor
from sklearn.svm import SVR
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.linear_model import ElasticNet
import numpy as np
from sklearn.neighbors import KNeighborsRegressor
from sklearn.feature_selection import SelectFromModel
import matplotlib.pyplot as plt
import lightgbm as lgb
from openpyxl import load_workbook
from sklearn.model_selection import KFold
from sklearn.model_selection import GridSearchCV
from sklearn.linear_model import Ridge
from sklearn.linear_model import LinearRegression
from keras.models import Sequential
from keras.layers import LSTM, Dense, Dropout
from keras.callbacks import EarlyStopping
from tensorflow.keras.optimizers import Adam
from sklearn.linear_model import Lasso
import openpyxl

def create_excel_file(filename='SA_CM.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active

    method_names = {
        'A1': 'Method Name',
        'A2': 'Artificial Neural Network',
        'A3': 'Decision Tree',
        'A4': 'Elastic Net Regression',
        'A5': 'General Regression Neural Network',
        'A6': 'K-Nearest Neighbors',
        'A7': 'Lasso Regression',
        'A8': 'LightGBM',
        'A9': 'Linear Regression',
        'A10': 'Long Short-Term Memory',
        'A11': 'Probabilistic Neural Network',
        'A12': 'Radial Basis Function',
        'A13': 'Random Forest',
        'A14': 'Ridge Regression',
        'A15': 'Support Vector Machine',
        'A16': 'XGBoost',
    }

    for cell, method in method_names.items():
        ws[cell] = method

    ws['B1'] = 'Training R2'
    ws['C1'] = 'Testing R2'
    ws['D1'] = 'Method R2'

    wb.save(filename)
    
    print("Exel File Has Been Created.")

create_excel_file()

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

model = XGBRegressor(objective='reg:squarederror', random_state=42, n_estimators=100, learning_rate=0.1)

model.fit(X_train, y_train)

y_train_pred = model.predict(X_train)
y_test_pred = model.predict(X_test)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test
predicted_data['Vs Predicted (Km/s)'] = y_test_pred
predicted_data.to_excel('Predicted_Well_B_XGBOOST.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B16'] = r2_train
sheet['C16'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()  
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

param_grid = {
    'C': [0.1, 1, 10, 100],
    'epsilon': [0.01, 0.1, 0.2],
    'gamma': ['scale', 'auto', 0.01, 0.1]
}

grid_search = GridSearchCV(SVR(kernel='rbf'), param_grid, scoring='r2', cv=5)

grid_search.fit(X_train, y_train)

best_model = grid_search.best_estimator_

y_train_pred = best_model.predict(X_train)
y_test_pred = best_model.predict(X_test)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test  
predicted_data['Vs Predicted (Km/s)'] = y_test_pred  
predicted_data.to_excel('Predicted_Well_B_SVM.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Best parameters: {grid_search.best_params_}")
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx' 
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B15'] = r2_train
sheet['C15'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()  
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

poly = PolynomialFeatures(degree=2)  
X_train_poly = poly.fit_transform(X_train)
X_test_poly = poly.transform(X_test)

ridge_model = Ridge(alpha=1.0)  
ridge_model.fit(X_train_poly, y_train)

y_train_pred = ridge_model.predict(X_train_poly)
y_test_pred = ridge_model.predict(X_test_poly)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test
predicted_data['Vs Predicted (Km/s)'] = y_test_pred
predicted_data.to_excel('Predicted_Well_B_RidgeRegression.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B14'] = r2_train
sheet['C14'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

model = RandomForestRegressor(n_estimators=100, random_state=42)

model.fit(X_train, y_train)

y_train_pred = model.predict(X_train)
y_test_pred = model.predict(X_test)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test
predicted_data['Vs Predicted (Km/s)'] = y_test_pred
predicted_data.to_excel('Predicted_Well_B_RF.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B13'] = r2_train
sheet['C13'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()  
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

def rbf_predict(X_train, y_train, X_test, sigma):
    predictions = []
    for x in X_test:
        
        distances = np.linalg.norm(X_train - x, axis=1)
        
        weights = np.exp(-(distances**2) / (2 * sigma**2))
        
        prediction = np.sum(weights * y_train) / np.sum(weights)
        predictions.append(prediction)
    return np.array(predictions)

sigma = 0.2  

y_train_pred = rbf_predict(X_train, y_train, X_train, sigma)
y_test_pred = rbf_predict(X_train, y_train, X_test, sigma)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test  
predicted_data['Vs Predicted (Km/s)'] = y_test_pred  
predicted_data.to_excel('Predicted_Well_B_RBF.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B12'] = r2_train
sheet['C12'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()  
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

def pnn_predict(X_train, y_train, X_test, sigma):
    predictions = []
    for x in X_test:
        
        distances = np.linalg.norm(X_train - x, axis=1)
        
        weights = np.exp(-(distances**2) / (2 * sigma**2))
        
        prediction = np.sum(weights * y_train) / np.sum(weights)
        predictions.append(prediction)
    return np.array(predictions)

sigma = 0.2  

y_train_pred = pnn_predict(X_train, y_train, X_train, sigma)
y_test_pred = pnn_predict(X_train, y_train, X_test, sigma)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test  
predicted_data['Vs Predicted (Km/s)'] = y_test_pred  
predicted_data.to_excel('Predicted_Well_B_PNN.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B11'] = r2_train
sheet['C11'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()  
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

X_train = X_train.reshape((X_train.shape[0], 1, X_train.shape[1]))
X_test = X_test.reshape((X_test.shape[0], 1, X_test.shape[1]))

split_index = int(0.8 * len(X_train))  
X_val = X_train[split_index:]
y_val = y_train[split_index:]
X_train = X_train[:split_index]
y_train = y_train[:split_index]

model = Sequential()
model.add(LSTM(100, activation='relu', return_sequences=True, input_shape=(X_train.shape[1], X_train.shape[2])))
model.add(Dropout(0.2))  
model.add(LSTM(50, activation='relu', return_sequences=False))  
model.add(Dropout(0.2))  
model.add(Dense(1))  
model.compile(optimizer='adam', loss='mean_squared_error')
early_stopping = EarlyStopping(monitor='val_loss', patience=15, restore_best_weights=True)
model.fit(X_train, y_train, epochs=200, batch_size=32, validation_data=(X_val, y_val), callbacks=[early_stopping], verbose=1)

y_train_pred = model.predict(X_train)
y_val_pred = model.predict(X_val)
y_test_pred = model.predict(X_test)

predicted_data = pd.DataFrame(X_test.reshape(X_test.shape[0], -1), columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test
predicted_data['Vs Predicted (Km/s)'] = y_test_pred
predicted_data.to_excel('Predicted_Well_B_LSTM.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_val = r2_score(y_val, y_val_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Validation R2 score: {r2_val:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred.flatten()
test_error = y_test - y_test_pred.flatten()

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B10'] = r2_train
sheet['C10'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()  
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

lr_model = LinearRegression()
lr_model.fit(X_train, y_train)

y_train_pred = lr_model.predict(X_train)
y_test_pred = lr_model.predict(X_test)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test
predicted_data['Vs Predicted (Km/s)'] = y_test_pred
predicted_data.to_excel('Predicted_Well_B_LinearRegression.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B09'] = r2_train
sheet['C09'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()  
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

poly = PolynomialFeatures(degree=2)  
X_train_poly = poly.fit_transform(X_train)
X_test_poly = poly.transform(X_test)

train_data = lgb.Dataset(X_train_poly, label=y_train)
test_data = lgb.Dataset(X_test_poly, label=y_test, reference=train_data)

params = {
    'objective': 'regression',
    'metric': 'rmse',
    'boosting_type': 'gbdt',
    'learning_rate': 0.1,
    'num_leaves': 31,
    'max_depth': -1,
    'verbose': -1,
}

lgb_model = lgb.train(
    params,
    train_data,
    num_boost_round=1000,
    valid_sets=[test_data],
    valid_names=['test'],
)

y_train_pred = lgb_model.predict(X_train_poly, num_iteration=lgb_model.best_iteration)
y_test_pred = lgb_model.predict(X_test_poly, num_iteration=lgb_model.best_iteration)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test
predicted_data['Vs Predicted (Km/s)'] = y_test_pred
predicted_data.to_excel('Predicted_Well_B_LightGBM.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B08'] = r2_train
sheet['C08'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()  
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

poly = PolynomialFeatures(degree=2)  
X_train_poly = poly.fit_transform(X_train)
X_test_poly = poly.transform(X_test)

param_grid = {
    'alpha': np.logspace(-4, 0, 50)  
}

lasso = Lasso()
grid_search = GridSearchCV(lasso, param_grid, cv=5, scoring='r2')
grid_search.fit(X_train_poly, y_train)

best_lasso_model = grid_search.best_estimator_

y_train_pred = best_lasso_model.predict(X_train_poly)
y_test_pred = best_lasso_model.predict(X_test_poly)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test
predicted_data['Vs Predicted (Km/s)'] = y_test_pred
predicted_data.to_excel('Predicted_Well_B_LassoRegression.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B07'] = r2_train
sheet['C07'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()  
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

rf = RandomForestRegressor(n_estimators=100, random_state=42)
rf.fit(X_train, y_train)

selector = SelectFromModel(rf, threshold='mean', prefit=True)
X_train_selected = selector.transform(X_train)
X_test_selected = selector.transform(X_test)

knn = KNeighborsRegressor()

param_grid = {
    'n_neighbors': np.arange(1, 10),
    'weights': ['uniform', 'distance'],
    'metric': ['euclidean', 'manhattan']
}

grid_search = GridSearchCV(knn, param_grid, cv=10, scoring='r2', n_jobs=-1)
grid_search.fit(X_train_selected, y_train)

best_knn = grid_search.best_estimator_

best_knn.fit(X_train_selected, y_train)

y_train_pred = best_knn.predict(X_train_selected)
y_test_pred = best_knn.predict(X_test_selected)

selected_feature_indices = selector.get_support(indices=True)
selected_feature_names = ['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']  
selected_feature_names = [selected_feature_names[i] for i in selected_feature_indices]

predicted_data = pd.DataFrame(X_test_selected, columns=selected_feature_names)  
predicted_data['Vs (Km/s)'] = y_test  
predicted_data['Vs Predicted (Km/s)'] = y_test_pred  
predicted_data.to_excel('Predicted_Well_B_KNN.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B06'] = r2_train
sheet['C06'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

def grnn_predict(X_train, y_train, X_test, sigma):
    predictions = []
    for x in X_test:
        distances = np.linalg.norm(X_train - x, axis=1)
        weights = np.exp(-(distances**2) / (2 * sigma**2))
        prediction = np.sum(weights * y_train) / np.sum(weights)
        predictions.append(prediction)
    return np.array(predictions)

def cross_validate_sigma(X, y, sigma_values, n_splits=5):
    best_sigma = None
    best_r2 = -np.inf
    kf = KFold(n_splits=n_splits)

    for sigma in sigma_values:
        r2_scores = []
        for train_index, val_index in kf.split(X):
            X_train_cv, X_val_cv = X[train_index], X[val_index]
            y_train_cv, y_val_cv = y[train_index], y[val_index]
            y_val_pred = grnn_predict(X_train_cv, y_train_cv, X_val_cv, sigma)
            r2 = r2_score(y_val_cv, y_val_pred)
            r2_scores.append(r2)
        
        mean_r2 = np.mean(r2_scores)
        if mean_r2 > best_r2:
            best_r2 = mean_r2
            best_sigma = sigma
            
    return best_sigma

sigma_values = np.linspace(0.03, 0.8, 20)  
best_sigma = cross_validate_sigma(X_train, y_train, sigma_values)
print(f"Best sigma from cross-validation: {best_sigma:.4f}")

y_train_pred = grnn_predict(X_train, y_train, X_train, best_sigma)
y_test_pred = grnn_predict(X_train, y_train, X_test, best_sigma)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test  
predicted_data['Vs Predicted (Km/s)'] = y_test_pred  
predicted_data.to_excel('Predicted_Well_B_GRNN.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B05'] = r2_train
sheet['C05'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

poly = PolynomialFeatures(degree=2)
X_train_poly = poly.fit_transform(X_train)
X_test_poly = poly.transform(X_test)

param_grid = {
    'alpha': np.logspace(-4, 0, 50),
    'l1_ratio': np.linspace(0, 1, 10)
}

elastic_net = ElasticNet()
grid_search = GridSearchCV(elastic_net, param_grid, cv=5, scoring='r2')
grid_search.fit(X_train_poly, y_train)

best_en_model = grid_search.best_estimator_

y_train_pred = best_en_model.predict(X_train_poly)
y_test_pred = best_en_model.predict(X_test_poly)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test
predicted_data['Vs Predicted (Km/s)'] = y_test_pred
predicted_data.to_excel('Predicted_Well_B_ElasticNetRegression.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B04'] = r2_train
sheet['C04'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

model = DecisionTreeRegressor(random_state=42)

model.fit(X_train, y_train)

y_train_pred = model.predict(X_train)
y_test_pred = model.predict(X_test)

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test
predicted_data['Vs Predicted (Km/s)'] = y_test_pred
predicted_data.to_excel('Predicted_Well_B_DT.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B03'] = r2_train
sheet['C03'] = r2_test

workbook.save(excel_file_path)

def load_and_preprocess_data(file_path):
    data = pd.read_excel(file_path, sheet_name='Sheet1')
    data = data.replace(' ', np.nan).dropna()
    X = data[['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)']].values.astype(float)
    y = data['Vs (Km/s)'].values.astype(float)
    return X, y

X_train, y_train = load_and_preprocess_data('Well A.xlsx')
X_test, y_test = load_and_preprocess_data('Well B.xlsx')

scaler_X = StandardScaler()
X_train = scaler_X.fit_transform(X_train)
X_test = scaler_X.transform(X_test)

ann_model = Sequential([
    Dense(64, input_shape=(X_train.shape[1],), activation='relu'),
    Dense(32, activation='relu'),
    Dense(1, activation='linear')
])

ann_model.compile(optimizer=Adam(learning_rate=0.001), loss='mean_squared_error')

early_stopping = EarlyStopping(monitor='val_loss', patience=10, restore_best_weights=True)

ann_model.fit(X_train, y_train, epochs=100, batch_size=32, validation_split=0.2, 
              callbacks=[early_stopping], verbose=1)

y_train_pred = ann_model.predict(X_train).flatten()
y_test_pred = ann_model.predict(X_test).flatten()

predicted_data = pd.DataFrame(X_test, columns=['RHOB (g/cc)', 'DTCO (us/ft)', 'GR (API)', 'NPHI (v/v)'])
predicted_data['Vs (Km/s)'] = y_test
predicted_data['Vs Predicted (Km/s)'] = y_test_pred
predicted_data.to_excel('Predicted_Well_B_ANN.xlsx', index=False)

r2_train = r2_score(y_train, y_train_pred)
r2_test = r2_score(y_test, y_test_pred)
print(f"Training R2 score: {r2_train:.4f}")
print(f"Test R2 score: {r2_test:.4f}")

train_error = y_train - y_train_pred
test_error = y_test - y_test_pred

excel_file_path = 'SA_CM.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active 

sheet['B02'] = r2_train
sheet['C02'] = r2_test

workbook.save(excel_file_path)

def calculate_and_save_r2_1():
    input_file = 'Predicted_Well_B_ANN.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D02'] = r2
    workbook.save(output_file)

calculate_and_save_r2_1()

def calculate_and_save_r2_2():
    input_file = 'Predicted_Well_B_DT.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D03'] = r2
    workbook.save(output_file)
    
calculate_and_save_r2_2()

def calculate_and_save_r2_3():
    input_file = 'Predicted_Well_B_ElasticNetRegression.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D04'] = r2
    workbook.save(output_file)

calculate_and_save_r2_3()

def calculate_and_save_r2_4():
    input_file = 'Predicted_Well_B_GRNN.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D05'] = r2
    workbook.save(output_file)

calculate_and_save_r2_4()

def calculate_and_save_r2_5():
    input_file = 'Predicted_Well_B_KNN.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 1].values
    y_predicted = data.iloc[2:2486, 2].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D06'] = r2
    workbook.save(output_file)

calculate_and_save_r2_5()

def calculate_and_save_r2_6():
    input_file = 'Predicted_Well_B_LassoRegression.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D07'] = r2
    workbook.save(output_file)

calculate_and_save_r2_6()

def calculate_and_save_r2_7():
    input_file = 'Predicted_Well_B_LightGBM.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D08'] = r2
    workbook.save(output_file)

calculate_and_save_r2_7()

def calculate_and_save_r2_8():
    input_file = 'Predicted_Well_B_LinearRegression.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D09'] = r2
    workbook.save(output_file)

calculate_and_save_r2_8()

def calculate_and_save_r2_9():
    input_file = 'Predicted_Well_B_LSTM.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D10'] = r2
    workbook.save(output_file)

calculate_and_save_r2_9()

def calculate_and_save_r2_10():
    input_file = 'Predicted_Well_B_PNN.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D11'] = r2
    workbook.save(output_file)

calculate_and_save_r2_10()

def calculate_and_save_r2_11():
    input_file = 'Predicted_Well_B_RBF.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D12'] = r2
    workbook.save(output_file)

calculate_and_save_r2_11()

def calculate_and_save_r2_12():
    input_file = 'Predicted_Well_B_RF.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D013'] = r2
    workbook.save(output_file)

calculate_and_save_r2_12()

def calculate_and_save_r2_13():
    input_file = 'Predicted_Well_B_RidgeRegression.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D014'] = r2
    workbook.save(output_file)

calculate_and_save_r2_13()

def calculate_and_save_r2_14():
    input_file = 'Predicted_Well_B_SVM.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D015'] = r2
    workbook.save(output_file)

calculate_and_save_r2_14()

def calculate_and_save_r2_15():
    input_file = 'Predicted_Well_B_XGBOOST.xlsx'
    output_file = "SA_CM.xlsx"
    data = pd.read_excel(input_file)
    y_actual = data.iloc[2:2486, 4].values
    y_predicted = data.iloc[2:2486, 5].values
    n = 2485
    y_mean = sum(y_actual) / n
    ss_res = sum((y_actual - y_predicted) ** 2)
    ss_tot = sum((y_actual - y_mean) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    workbook = load_workbook(output_file)
    sheet = workbook.active
    sheet['D016'] = r2
    workbook.save(output_file)

calculate_and_save_r2_15()

print(f"Finished!")